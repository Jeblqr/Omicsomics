"""
Data export utilities.

Exports unified format data to various formats (CSV, TSV, Excel, JSON).
"""

import csv
import io
import json
from typing import Dict, List, Optional, Any

from app.schemas.unified import UnifiedData


class DataExporter:
    """Export unified data to various formats."""

    @staticmethod
    def to_csv(
        unified_data: UnifiedData,
        include_columns: Optional[List[str]] = None,
        delimiter: str = ",",
    ) -> bytes:
        """
        Export unified data to CSV format.

        Args:
            unified_data: Unified data object
            include_columns: List of columns to include (None = all)
            delimiter: Field delimiter (default: comma)

        Returns:
            CSV data as bytes
        """
        output = io.StringIO()

        # Determine columns
        if not unified_data.features:
            raise ValueError("No features to export")

        sample_feature = unified_data.features[0]
        all_columns = list(sample_feature.model_dump().keys())

        columns = include_columns if include_columns else all_columns

        # Write CSV
        writer = csv.DictWriter(output, fieldnames=columns, delimiter=delimiter)
        writer.writeheader()

        for feature in unified_data.features:
            feature_dict = feature.model_dump()
            row = {col: feature_dict.get(col, "") for col in columns}

            # Convert lists/dicts to JSON strings
            for key, value in row.items():
                if isinstance(value, (list, dict)):
                    row[key] = json.dumps(value)

            writer.writerow(row)

        return output.getvalue().encode("utf-8")

    @staticmethod
    def to_tsv(
        unified_data: UnifiedData,
        include_columns: Optional[List[str]] = None,
    ) -> bytes:
        """
        Export unified data to TSV format.

        Args:
            unified_data: Unified data object
            include_columns: List of columns to include (None = all)

        Returns:
            TSV data as bytes
        """
        return DataExporter.to_csv(unified_data, include_columns, delimiter="\t")

    @staticmethod
    def to_excel(
        unified_data: UnifiedData,
        include_columns: Optional[List[str]] = None,
        sheet_name: str = "Data",
    ) -> bytes:
        """
        Export unified data to Excel format.

        Args:
            unified_data: Unified data object
            include_columns: List of columns to include (None = all)
            sheet_name: Name of the Excel sheet

        Returns:
            Excel data as bytes
        """
        try:
            import openpyxl
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise ImportError(
                "openpyxl is required for Excel export. Install with: pip install openpyxl"
            )

        if not unified_data.features:
            raise ValueError("No features to export")

        # Determine columns
        sample_feature = unified_data.features[0]
        all_columns = list(sample_feature.model_dump().keys())
        columns = include_columns if include_columns else all_columns

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name

        # Write header
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = openpyxl.styles.Font(bold=True)

        # Write data
        for row_idx, feature in enumerate(unified_data.features, 2):
            feature_dict = feature.model_dump()
            for col_idx, col_name in enumerate(columns, 1):
                value = feature_dict.get(col_name, "")

                # Convert lists/dicts to JSON strings
                if isinstance(value, (list, dict)):
                    value = json.dumps(value)

                ws.cell(row=row_idx, column=col_idx, value=value)

        # Auto-adjust column widths
        for col_idx, col_name in enumerate(columns, 1):
            column_letter = get_column_letter(col_idx)
            max_length = len(col_name)

            for row_idx in range(2, min(102, len(unified_data.features) + 2)):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))

            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()

    @staticmethod
    def to_json(
        unified_data: UnifiedData,
        include_columns: Optional[List[str]] = None,
        pretty: bool = True,
    ) -> bytes:
        """
        Export unified data to JSON format.

        Args:
            unified_data: Unified data object
            include_columns: List of columns to include (None = all)
            pretty: Pretty-print JSON (default: True)

        Returns:
            JSON data as bytes
        """
        if not unified_data.features:
            data = {
                "metadata": unified_data.metadata.model_dump(),
                "features": [],
            }
        else:
            # Determine columns
            sample_feature = unified_data.features[0]
            all_columns = list(sample_feature.model_dump().keys())
            columns = include_columns if include_columns else all_columns

            # Convert features to list of dicts
            features_list = []
            for feature in unified_data.features:
                feature_dict = feature.model_dump()
                filtered_dict = {
                    col: feature_dict.get(col) for col in columns if col in feature_dict
                }
                features_list.append(filtered_dict)

            data = {
                "metadata": unified_data.metadata.model_dump(),
                "features": features_list,
            }

        if pretty:
            json_str = json.dumps(data, indent=2, ensure_ascii=False)
        else:
            json_str = json.dumps(data, ensure_ascii=False)

        return json_str.encode("utf-8")

    @staticmethod
    def export(
        unified_data: UnifiedData,
        format: str,
        include_columns: Optional[List[str]] = None,
        **kwargs,
    ) -> bytes:
        """
        Export unified data to specified format.

        Args:
            unified_data: Unified data object
            format: Export format ('csv', 'tsv', 'excel', 'json')
            include_columns: List of columns to include (None = all)
            **kwargs: Additional format-specific arguments

        Returns:
            Exported data as bytes

        Raises:
            ValueError: If format is not supported
        """
        format_lower = format.lower()

        if format_lower == "csv":
            return DataExporter.to_csv(unified_data, include_columns)
        elif format_lower == "tsv":
            return DataExporter.to_tsv(unified_data, include_columns)
        elif format_lower in ["excel", "xlsx"]:
            return DataExporter.to_excel(unified_data, include_columns, **kwargs)
        elif format_lower == "json":
            return DataExporter.to_json(unified_data, include_columns, **kwargs)
        else:
            raise ValueError(f"Unsupported export format: {format}")

    @staticmethod
    def get_mime_type(format: str) -> str:
        """Get MIME type for export format."""
        format_lower = format.lower()

        mime_types = {
            "csv": "text/csv",
            "tsv": "text/tab-separated-values",
            "excel": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "json": "application/json",
        }

        return mime_types.get(format_lower, "application/octet-stream")

    @staticmethod
    def get_file_extension(format: str) -> str:
        """Get file extension for export format."""
        format_lower = format.lower()

        extensions = {
            "csv": "csv",
            "tsv": "tsv",
            "excel": "xlsx",
            "xlsx": "xlsx",
            "json": "json",
        }

        return extensions.get(format_lower, "dat")
